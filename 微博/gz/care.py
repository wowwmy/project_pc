import requests
import time
from head import headers
import openpyxl
import logging


class WeiBo:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        ws = self.wb['Sheet']
        self.wb.remove(ws)

    def __del__(self):
        self.wb.save('微博.xlsx')

    # 获取账户下的所有频道id
    def get_id(self):
        url = 'https://weibo.com/ajax/feed/allGroups'
        params = {
            "is_new_segment": "1",
            "fetch_hot": "1"
        }
        res = requests.get(url=url, params=params, headers=headers)
        return res.json()

    # 获取 该账户下的最新微博频道的 id
    def get_all_care_id(self, res):
        groups = res['groups']
        gid = ''
        for group in groups:
            if group['title'] == '默认分组':
                # 获取全部关注
                for g in group['group']:
                    if g['title'] == '最新微博':
                        gid = g['gid']
                        break
                break
        return gid

    # 获取文章的长(全部)文本
    def all_text(self, mblogid):
        url = 'https://weibo.com/ajax/statuses/longtext'
        res = requests.get(url=url + f'?id={mblogid}', headers=headers).json()
        if res['data'] == {}:
            return None
        text = res['data']['longTextContent']
        return text

    # 获取 最新微博(你关注博主的所有文章)下的文章信息
    def get_essay(self, ids):
        n = 0
        url = 'https://weibo.com/ajax/feed/friendstimeline'
        params = {
            "list_id": f"{ids}",
            "refresh": "4",
            "since_id": "0",
            "count": "25",
            "fid": f"{ids}"
        }
        res = requests.get(url=url, headers=headers, params=params).json()
        while True:
            # 下一页信息
            max_id = res['max_id']
            # 文章数据
            statuses = res['statuses']
            for s in statuses:
                user = s['user']['screen_name']  # 用户名
                mblogid = s['mblogid']
                text = self.all_text(mblogid)  # 文本
                if text:
                    pass
                else:
                    text = s['text_raw']
                creat_time = s['created_at']  # 创建时间
                try:
                    region_name = s['region_name'][4:]  # 发布地点
                except KeyError:
                    region_name = ''
                support = s['attitudes_count']  # 点赞
                reply_count = s['comments_count']  # 回复数
                repeat_count = s['reposts_count']  # 转发数
                self.save([user, creat_time, region_name, support, reply_count, repeat_count, text])
            n += 1
            print(f'请求成功   {n}     页')
            if max_id != 0:
                params = {
                    "list_id": f"{ids}",
                    "refresh": "4",
                    "max_id": max_id,
                    "count": "25",
                    "fid": f"{ids}"
                }
                res = requests.get(url=url, headers=headers, params=params).json()
            else:
                break

    # 保存
    def save(self, data: list):
        sheets = self.wb.sheetnames
        if data[0] not in sheets:
            ws = self.wb.create_sheet(title=data[0])
            ws.append(['作者', '创建时间', '发布地点', '点赞数', '回复数', '转发数', '文本'])
            ws.append(data)
        else:
            ws = self.wb[data[0]]
            ws.append(data)

    def run(self):
        res = self.get_id()
        ids = self.get_all_care_id(res)
        self.get_essay(ids)


if __name__ == '__main__':
    WeiBo().run()
