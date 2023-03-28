import requests
import time
import re, json
import openpyxl
import execjs


class Config:
    def __init__(self):
        self.headers = {
            "referer": "https://www.autohome.com.cn/",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"
        }
        self.wb = openpyxl.Workbook()
        ws = self.wb['Sheet']
        self.wb.remove(ws)
        self.ws = self.wb.create_sheet('汽车')

    def __del__(self):
        self.wb.save('汽车之家.xlsx')

    # 请求
    def request(self, url):
        n = 0
        while True:
            try:
                res = requests.get(url=url, headers=self.headers)
                if res.status_code == 200:
                    break
            except requests.exceptions.ConnectTimeout:
                if n < 5:
                    n += 1
                    time.sleep(10)
                else:
                    res = None
                    break
        return res

    # 保存
    def save(self, data):
        for d in data:
            self.ws.append(d)

    # 提取数值，将数值处理成对应的条数据
    def draw(self, data):
        data_list = []
        c_len = len(data[0])  # 统计多少台车
        for i in range(c_len):
            v = []
            for d in data:
                v.append(d[i])
            data_list.append(v)
        return data_list

    # 添加表格头部标签
    def add_sheet(self, data):
        if '汽车' in self.wb.sheetnames:
            pass
        else:
            self.ws.append(data)

    # 解析
    def parse(self, res, font_list):
        # keyLink,option,bag    三个参数对象三个数据,更加需要来获取,修改代码
        # keyLink = re.findall(r'var keyLink = (.*?);', res)[0]
        config = re.findall(r'var config = (.*?);', res)[0]  # 基本参数配置
        # option = re.findall(r'var option = (.*?);', res)[0]
        # bag = re.findall(r'var bag = (.*?);', res)[0]
        paramtypeitems = json.loads(config)['result']['paramtypeitems']
        names = []  # 存储名称
        values = []  # 存储数值
        for p in paramtypeitems:
            name, value = self.parse_detail(p['paramitems'], font_list)
            names.extend(name)
            values.extend(value)
        self.add_sheet(names)
        v = self.draw(values)
        self.save(v)

    # 具体数据
    def parse_detail(self, items, font_list):
        names = []  # 名称
        config_values = []  # 基础参数值
        for item in items:
            v = []
            name = self.filter(item['name'], font_list)
            names.append(name)
            for value in item['valueitems']:
                valueitems = self.filter(value['value'], font_list)
                v.append(valueitems)
            config_values.append(v)
        return names, config_values

    # 过滤替换掉标签变成字体
    def filter(self, value, font_list):
        # 提取 class类属性中的 数字
        spans = re.findall(r"<span(.*?)</span>", value)
        for s in spans:
            # 匹配出第一个数字
            num = re.search(r"hs_kw([\d]+)_", value).group(1)
            # 将 <span ... </span> 替换成文字
            value = re.sub(fr"<span{s}</span>", font_list[int(num)], value)
        return value

    # 获取字体js
    def get_font_js(self, html):
        script = re.findall(r'<script>\(function\((.*?)document\);', html)
        font_list = []
        for s in script:
            if 'config' in s:
                js_txt = '''
                const jsdom = require('jsdom');
                const { JSDOM } = jsdom;
                const dom = new JSDOM('<!DOCTYPE html><p>hello world</p>');
                window=dom.window
                document=window.document
                navigator=window.navigator;
                var _dict = [];
                '''
                return_js = 'function get_dict(){return _dict}'
                new_s = js_txt + '(function(' + s + 'document);' + return_js
                new_js = re.sub(r'\$InsertRule\$\(\$index\$, \$temp\$\);',
                                '_dict.push($temp$);$InsertRule$($index$, $temp$);', new_s)
                font_list = execjs.compile(new_js).call('get_dict')
                break
        return font_list

    def run(self, urls):
        for url in urls:
            url = f'https://car.autohome.com.cn/config/series/{url}.html'
            res = self.request(url)
            # 获取反爬字体列表
            font_list = self.get_font_js(res.text)
            if res:
                self.parse(res.text, font_list)
            else:
                pass


if __name__ == '__main__':
    # 输入品牌id 即.html前面的数字
    Config().run(['5769', '6899'])
